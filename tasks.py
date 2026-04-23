from robocorp.tasks import task
from robocorp import browser
import os
import tkinter as tk
from tkinter import simpledialog
import oracledb
import subprocess
import tempfile
import glob
import openpyxl

def conectar_bd():
    print("\n=== Conectando a Base de Datos Oracle (Modo Grueso) ===")
    try:
        oracledb.init_oracle_client(lib_dir=r"C:\oracle\instantclient_23_0")
        conn = oracledb.connect(
            user="datasoft",
            password="data2001",
            dsn="172.16.14.17:1521/ORCL"
        )
        print("Conexión exitosa a la BD.")
        return conn
    except Exception as e:
        print(f"Error al conectar a BD: {e}")
        return None

def ejecutar_consulta_db(query, conn):
    if not conn:
        print("Aviso: No hay conexión a BD para ejecutar la consulta.")
        return "0"
    try:
        with conn.cursor() as cursor:
            cursor.execute(query)
            res = cursor.fetchone()
            if res:
                return str(res[0])
            return "0"
    except Exception as e:
        print(f"Error ejecutando consulta en BD: {e}")
        return "0"

def ejecutar_actualizacion_db(query, conn):
    if not conn:
        print("Aviso: No hay conexión a BD para ejecutar la actualización.")
        return False
    try:
        with conn.cursor() as cursor:
            cursor.execute(query)
            conn.commit()
            return True
    except Exception as e:
        print(f"Error ejecutando actualización en BD: {e}")
        return False

def ejecutar_consulta_fila_db(query, conn):
    """Retorna la primera fila completa como dict o None."""
    if not conn:
        return None
    try:
        with conn.cursor() as cursor:
            cursor.execute(query)
            row = cursor.fetchone()
            if row and cursor.description:
                cols = [col[0] for col in cursor.description]
                return dict(zip(cols, row))
            return None
    except Exception as e:
        print(f"Error ejecutando consulta de fila en BD: {e}")
        return None

# Mapeo de Línea de Negocio (Helix) → (CMPN_CODIGO, ESOR_CODIGO)
LINEAS_NEGOCIO = {
    "bancolombia": [("BANCOLOMBI", "PIC")],
    "sufi":        [("SUFI",       "PIC")],
    "leasing":     [("LEASINGBAN", "ARCHIVO")],
    "factoring":   [("FACTBANCOL", "GARANTIAS")],
    "todos":       [
        ("BANCOLOMBI", "PIC"),
        ("LEASINGBAN", "ARCHIVO"),
        ("SUFI",       "PIC"),
        ("FACTBANCOL", "GARANTIAS"),
    ],
}

def lineas_desde_helix(linea_negocio_texto):
    """Convierte el texto de Línea de Negocio de Helix a lista de (CMPN, ESOR)."""
    texto = linea_negocio_texto.strip().lower()
    for clave, valor in LINEAS_NEGOCIO.items():
        if clave in texto:
            return valor
    return [("BANCOLOMBI", "PIC")]  # Default

def generar_login(nombre_completo, conn):
    """
    Genera un LOGIN tomando las 2 primeras letras de cada palabra,
    hasta un máximo de 4 palabras (8 caracteres). Si el LOGIN ya existe,
    agrega un dígito incremental al final.
    """
    palabras = nombre_completo.upper().split()
    palabras = palabras[:4]  # máximo 4 palabras
    base = "".join([p[:2] for p in palabras if len(p) >= 2])
    login = base
    sufijo = 1
    while True:
        res = ejecutar_consulta_db(f"SELECT COUNT(*) FROM a_usuario WHERE LOGIN = '{login}'", conn)
        if res and res.strip() == "0":
            return login
        login = base + str(sufijo)
        sufijo += 1

def generar_informe_glpi(page, datos_extraidos=None, login_generado=None, password_generado=None, estado_helix=None, titulo_glpi="ACTIVACION USUARIO BANCO", id_ticket=None):
    print("\n=== Iniciando creación de ticket en GLPI ===")
    try:
        context = page.context
        glpi_page = context.new_page()
        
        glpi_page.goto("http://172.0.0.60/Mesa_Servicios_Tecnologicos/index.php")
        
        # Comprobar si pide login o si ya hay sesión guardada en memoria
        try:
            # Esperamos 3 segundos a ver si aparece el campo de usuario
            glpi_page.locator("#login_name").wait_for(state="visible", timeout=3000)
            necesita_login = True
        except:
            necesita_login = False

        if necesita_login:
            print("Ingresando credenciales en GLPI...")
            glpi_page.locator("#login_name").fill("santiago.pinerez")
            # Selector genérico para evitar fallos por names dinámicos
            glpi_page.locator("input[type='password']").fill("Spadata55/")
            glpi_page.locator("[name=submit]").click()
            # Espera breve para asegurar que el inicio de sesión termine
            glpi_page.wait_for_timeout(2000)
        else:
            print("Sesión de GLPI detectada en caché. Omitiendo login...")
        
        print("Navegando al menú Crear caso...")
        glpi_page.locator("a", has_text="Soporte").click()
        glpi_page.locator("a", has_text="Crear caso").click()
        
        print("Esperando la carga del formulario de ticket...")
        # 1. Seleccionar Categoría (id dinámico)
        categoria_selector = "[id^='select2-dropdown_itilcategories_id']"
        glpi_page.locator(categoria_selector).wait_for(state="visible", timeout=15000)
        glpi_page.locator(categoria_selector).click()
        
        search_field = glpi_page.locator(".select2-search--dropdown > .select2-search__field")
        search_field.wait_for(state="visible", timeout=5000)
        search_field.fill("Aplicaciones > Administración de usuarios Bancolombia Banco")
        
        glpi_page.locator("li.select2-results__option", has_text="Aplicaciones > Administración de usuarios Bancolombia Banco").first.click()
        
        # 2. Asignado a
        print("Asignando usuario: Jose Luis Echavarria Ochoa...")
        
        # Esperamos a que apareza al menos un input de select2
        glpi_page.locator("input.select2-search__field").first.wait_for(state="visible", timeout=10000)
        
        # En el formulario estándar de GLPI hay 3 campos de actores: Solicitante, Observador y Asignado.
        # Seleccionamos explícitamente el tercero.
        inputs = glpi_page.locator("input.select2-search__field")
        count = inputs.count()
        
        if count >= 3:
            search_assign = inputs.nth(2)
            print("Campo de asignación identificado (Tercer select2).")
        else:
            search_assign = inputs.last
            print(f"Fallback: Utilizando el último campo select2 de {count} disponibles.")
            
        search_assign.scroll_into_view_if_needed()
        search_assign.click()
        search_assign.fill("Jose")
        
        # Esperamos que cargue la lista flotante y damos clic
        glpi_page.wait_for_timeout(2000) # dar un instante para que reaccione GLPI
        opcion_busqueda = glpi_page.locator("li.select2-results__option", has_text="Jose Luis Echavarria Ochoa").first
        opcion_busqueda.wait_for(state="visible", timeout=10000)
        opcion_busqueda.click()
        
        # 3. Título (id dinámico)
        print("Ingresando título...")
        titulo_input = glpi_page.locator("input[id^='name_']")
        titulo_input.fill(titulo_glpi)
        
        # 4. Descripción (dentro del iframe de TinyMCE)
        print("Ingresando descripción...")
        iframe_locator = glpi_page.frame_locator("iframe.tox-edit-area__iframe")
        body_locator = iframe_locator.locator("body#tinymce")
        body_locator.wait_for(state="visible", timeout=10000)
        
        # Construir la descripción de manera dinámica
        descripcion_glpi = f"Buen día,\nMe colaboran gestionando este ticket del Helix ({id_ticket if id_ticket else ''}) para activación de usuario de Bancolombia.\nadjunto evidencia:\n\n"
        
        if id_ticket:
            descripcion_glpi += f"ID Item Helix: {id_ticket}\n"
        
        if estado_helix:
            descripcion_glpi += f"Estado fijado en Helix: {estado_helix}\n\n"
        
        if datos_extraidos:
            descripcion_glpi += "--- Datos del usuario de Helix ---\n"
            for clave, valor in datos_extraidos.items():
                descripcion_glpi += f"{clave} {valor}\n"
                
        if login_generado and password_generado:
            descripcion_glpi += f"\n--- Usuario generado ---\nUsuario: {login_generado}\nContraseña: {password_generado}\n"
        
        body_locator.click()
        body_locator.fill(descripcion_glpi)
        # Presionamos un espacio y lo borramos para forzar a TinyMCE a registrar el evento de teclado
        # Esto elimina el estado nativo de error 'required' (rojo inactivo)
        body_locator.press("Space")
        body_locator.press("Backspace")
        
        # 5. Guardar (Agregar)
        print("Guardando caso en GLPI...")
        glpi_page.locator("[name=add]").click()
        
        # Dar tiempo al sistema para crear
        glpi_page.wait_for_timeout(3000)
        print("Caso creado en GLPI exitosamente.")
        
        # Se cierra la pestaña para no estorbar a Helix
        glpi_page.close()
    except Exception as e:
        print(f"Aviso: Error al intentar generar informe en GLPI: {e}")
        try:
            glpi_page.close() # Limpieza en caso de error
        except:
            pass

def descargar_adjunto_excel(page, frame):
    """
    Descarga el archivo Excel adjunto del ticket en Helix.
    Busca el enlace del adjunto dentro del iframe del ticket,
    hace clic para descargarlo y retorna la ruta del archivo descargado.
    """
    print("\n=== Descargando Excel adjunto del ticket ===")
    try:
        # Buscar el enlace del archivo adjunto dentro del frame
        # Los adjuntos en Helix suelen estar en un área de attachments
        adjunto_selector = (
            'a[href*=".xlsx"], a[href*=".xls"], '
            'a[title*=".xlsx"], a[title*=".xls"], '
            'span:has-text(".xlsx"), span:has-text(".xls"), '
            'button:has-text(".xlsx"), button:has-text(".xls")'
        )
        
        # Intentar primero en el frame, luego en la página
        loc = frame.locator(adjunto_selector).first if frame else page.locator(adjunto_selector).first
        
        # Si no se encuentra con selectores genéricos, buscar por el ícono de adjunto
        try:
            loc.wait_for(state="visible", timeout=5000)
        except:
            print("Buscando adjunto con selectores alternativos...")
            alt_selector = (
                '[class*="attachment"] a, '
                '[class*="Attachment"] a, '
                '[id*="attachment"] a, '
                'a[class*="file"], '
                'img[class*="attachment-icon"]'
            )
            loc = frame.locator(alt_selector).first if frame else page.locator(alt_selector).first
            loc.wait_for(state="visible", timeout=10000)
        
        # Configurar directorio de descarga temporal
        download_dir = os.path.join(tempfile.gettempdir(), "rpa_helix_downloads")
        os.makedirs(download_dir, exist_ok=True)
        
        # Iniciar la descarga haciendo clic
        with page.expect_download(timeout=30000) as download_info:
            loc.click()
        
        download = download_info.value
        archivo_descargado = os.path.join(download_dir, download.suggested_filename)
        download.save_as(archivo_descargado)
        
        print(f"Archivo descargado: {archivo_descargado}")
        return archivo_descargado
        
    except Exception as e:
        print(f"Error al descargar el adjunto Excel: {e}")
        # Fallback: buscar si ya hay un archivo Excel reciente en la carpeta de descargas del usuario
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        archivos_excel = glob.glob(os.path.join(downloads_path, "Planilla*.xlsx"))
        if archivos_excel:
            mas_reciente = max(archivos_excel, key=os.path.getmtime)
            print(f"Fallback: Usando archivo más reciente encontrado: {mas_reciente}")
            return mas_reciente
        return None

def leer_excel_masivo(ruta_archivo):
    """
    Lee el archivo Excel de solicitud masiva.
    Formato esperado:
      Fila 12: Encabezados (CEDULA, NOMBRE, CARGO, ÁREA, CORREO ELECTRONICO, USUARIO REFERENCIA)
      Columnas: B=Cédula, C=Nombre, D=Cargo, E=Área, F=Correo, G=Usuario referencia
      Desde fila 13 en adelante: datos de usuarios
    Retorna una lista de diccionarios con los datos de cada usuario.
    """
    print(f"\n=== Leyendo Excel masivo: {ruta_archivo} ===")
    usuarios = []
    
    try:
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        ws = wb.active  # Tomar la primera hoja
        
        # Verificar encabezados en fila 12
        encabezados = {
            'B': ws['B12'].value,
            'C': ws['C12'].value,
            'D': ws['D12'].value,
            'E': ws['E12'].value,
            'F': ws['F12'].value,
            'G': ws['G12'].value,
        }
        print(f"Encabezados detectados: {encabezados}")
        
        # Leer datos desde fila 13 en adelante
        fila = 13
        while True:
            cedula = ws[f'B{fila}'].value
            
            # Si la cédula está vacía, terminamos de leer
            if cedula is None or str(cedula).strip() == "":
                break
            
            nombre = ws[f'C{fila}'].value or ""
            correo = ws[f'F{fila}'].value or ""
            usuario_ref = ws[f'G{fila}'].value or ""
            
            usuario = {
                "cedula": str(cedula).strip(),
                "nombre": str(nombre).strip(),
                "correo": str(correo).strip(),
                "usuario_referencia": str(usuario_ref).strip(),
            }
            usuarios.append(usuario)
            fila += 1
        
        wb.close()
        print(f"Se leyeron {len(usuarios)} usuario(s) del Excel.")
        
    except Exception as e:
        print(f"Error al leer el Excel: {e}")
    
    return usuarios

def validar_fila_activacion(cedula, usuario_ref, conn):
    """
    Aplica la lógica de activación/desbloqueo para un usuario individual.
    En masiva, usuario_ref es el login del propio usuario a activar.
    Se busca primero por ese login y luego por cédula.
    Retorna un dict con el resultado: {exito: bool, login: str, mensaje: str}
    """
    print(f"\n  >> Validando activación para cédula: {cedula}, usuario: {usuario_ref}")
    
    existe = False
    login_final = usuario_ref
    
    # 1. Buscar primero por el login del usuario (columna G = login del propio usuario)
    if usuario_ref:
        res = ejecutar_consulta_db(
            f"SELECT COUNT(*) FROM a_usuario WHERE LOGIN = '{usuario_ref}'", conn
        )
        if res and res.strip() != "0":
            existe = True
            print(f"  Usuario encontrado por login: {login_final}")
    
    # 2. Si no se encontró por login, buscar por cédula
    if not existe and cedula:
        res_login = ejecutar_consulta_db(
            f"SELECT LOGIN FROM a_usuario WHERE NO_IDENTIFICACION = '{cedula}'", conn
        )
        if res_login and res_login.strip() not in ("0", ""):
            existe = True
            login_final = res_login.strip()
            print(f"  Usuario encontrado por cédula. LOGIN en BD: {login_final}")
    
    if existe and login_final:
        update_query = (
            f"UPDATE a_usuario SET CONTRASENA = NO_IDENTIFICACION, ESTADO = 'A', "
            f"CAMBIO_CONTRASENA = 'S' WHERE LOGIN = '{login_final}'"
        )
        exito = ejecutar_actualizacion_db(update_query, conn)
        
        if exito:
            print(f"  Activación exitosa para {login_final}")
            return {
                "exito": True,
                "login": login_final,
                "mensaje": f"Cédula {cedula} - Usuario {login_final} - Contraseña {cedula}: Activado exitosamente."
            }
        else:
            return {
                "exito": False,
                "login": login_final,
                "mensaje": f"Cédula {cedula} - Usuario {login_final}: Error al actualizar."
            }
    else:
        print(f"  Usuario con cédula {cedula} / login {usuario_ref} no existe en BD.")
        return {
            "exito": False,
            "login": None,
            "mensaje": f"Cédula {cedula} - Usuario {usuario_ref}: No existe el usuario. Requiere creación."
        }

def validar_fila_creacion(cedula, nombre_completo, email_usuario, usuario_ref, conn):
    """
    Aplica la lógica de creación para un usuario individual (masivo).
    Retorna un dict con el resultado: {exito: bool, login: str, mensaje: str}
    """
    print(f"\n  >> Validando creación para cédula: {cedula}, nombre: {nombre_completo}")
    
    # 1. Validar que el usuario de referencia exista
    ref_fila = ejecutar_consulta_fila_db(
        f"SELECT LOGIN, CMPN_CODIGO, ESOR_CODIGO, NOMBRE_USUARIO FROM a_usuario WHERE LOGIN = '{usuario_ref}'",
        conn
    )
    if not ref_fila:
        print(f"  Usuario de referencia '{usuario_ref}' no existe.")
        return {
            "exito": False,
            "login": None,
            "mensaje": f"Usuario de referencia '{usuario_ref}' no existe."
        }
    
    ref_cmpn = ref_fila.get("CMPN_CODIGO", "BANCOLOMBI")
    ref_esor = ref_fila.get("ESOR_CODIGO", "PIC")
    
    # 2. Verificar si usuario ya existe por cédula
    login_existente = ejecutar_consulta_db(
        f"SELECT LOGIN FROM a_usuario WHERE NO_IDENTIFICACION = '{cedula}'", conn
    )
    usuario_ya_existe = login_existente and login_existente.strip() not in ("0", "")
    
    # Para masiva usamos la línea del usuario de referencia
    lineas_requeridas = [(ref_cmpn, ref_esor)]
    
    if usuario_ya_existe:
        login_existente = login_existente.strip()
        print(f"  Usuario ya existe: {login_existente}. Verificando autorizaciones...")
        
        # Verificar qué líneas ya tiene en AUTORIZADO
        lineas_faltantes = []
        for cmpn, esor in lineas_requeridas:
            res_autr = ejecutar_consulta_db(
                f"SELECT COUNT(*) FROM AUTORIZADO WHERE AUTR_CODIGO = '{login_existente}' AND AUTR_CMPN_CODIGO = '{cmpn}'",
                conn
            )
            if res_autr and res_autr.strip() == "0":
                lineas_faltantes.append((cmpn, esor))
        
        if not lineas_faltantes:
            print(f"  El usuario {login_existente} ya tiene todas las autorizaciones.")
            return {
                "exito": False,
                "login": login_existente,
                "mensaje": f"Cédula {cedula} - Usuario {login_existente}: Ya existe con todas las autorizaciones."
            }
        else:
            # Añadir autorizaciones faltantes
            print(f"  Añadiendo {len(lineas_faltantes)} autorización(es) faltante(s)...")
            for cmpn, esor in lineas_faltantes:
                q_autr = (
                    f"INSERT INTO AUTORIZADO (AUTR_CMPN_CODIGO, AUTR_ESOR_CODIGO, AUTR_CODIGO, AUTR_NOMBRE) "
                    f"VALUES ('{cmpn}', '{esor}', '{login_existente}', '{nombre_completo}')"
                )
                ejecutar_actualizacion_db(q_autr, conn)
                q_serie = (
                    f"INSERT INTO autorizado_serie (AUSR_CMPN_CODIGO, AUSR_ESOR_CODIGO, AUSR_SRDC_CODIGO, AUSR_DCMT_CODIGO, AUSR_AUTR_CODIGO, AUTR_NOMBRE) "
                    f"SELECT a.AUSR_CMPN_CODIGO, a.AUSR_ESOR_CODIGO, a.AUSR_SRDC_CODIGO, a.AUSR_DCMT_CODIGO, '{login_existente}', '{nombre_completo}' "
                    f"FROM autorizado_serie a WHERE a.AUSR_AUTR_CODIGO = '{usuario_ref}' "
                    f"AND NOT EXISTS (SELECT 1 FROM autorizado_serie b WHERE b.AUSR_CMPN_CODIGO = a.AUSR_CMPN_CODIGO "
                    f"AND b.AUSR_ESOR_CODIGO = a.AUSR_ESOR_CODIGO "
                    f"AND b.AUSR_SRDC_CODIGO = a.AUSR_SRDC_CODIGO "
                    f"AND b.AUSR_DCMT_CODIGO = a.AUSR_DCMT_CODIGO "
                    f"AND b.AUSR_AUTR_CODIGO = '{login_existente}')"
                )
                ejecutar_actualizacion_db(q_serie, conn)
            
            return {
                "exito": True,
                "login": login_existente,
                "mensaje": f"Cédula {cedula} - Usuario {login_existente}: Autorizaciones añadidas exitosamente."
            }
    else:
        # Usuario no existe → Creación completa
        print(f"  Usuario no existe. Creando...")
        nuevo_login = generar_login(nombre_completo, conn)
        print(f"  LOGIN generado: {nuevo_login}")
        
        # 1. INSERT a_usuario
        q_usuario = (
            f"INSERT INTO a_usuario (ID_USUARIO, LOGIN, CONTRASENA, ID_GRUPO_USUARIO, NOMBRE_USUARIO, "
            f"ESTADO, LOGIN_USUARIO, FECHA_ULTIMA_ACT, CMPN_CODIGO, ESOR_CODIGO, E_MAIL, "
            f"TIPO_USUARIO, CAMBIO_CONTRASENA, FECHA_CAMBIO_CONTRASENA, NO_INTENTOS, "
            f"RESTABLECE_CONTRASENA, NO_IDENTIFICACION, FECHA_CREACION) "
            f"VALUES (seq_a_usuario.nextval, '{nuevo_login}', '{cedula}', 462, "
            f"'{nombre_completo}', 'A', 'SPINE', SYSDATE, '{ref_cmpn}', '{ref_esor}', "
            f"'{email_usuario}', 'E', 'N', SYSDATE, 0, 'S', '{cedula}', SYSDATE)"
        )
        exito_usuario = ejecutar_actualizacion_db(q_usuario, conn)
        if not exito_usuario:
            return {
                "exito": False,
                "login": None,
                "mensaje": f"Cédula {cedula}: Error al crear usuario ."
            }
        
        # 2. INSERT AUTORIZADO
        for cmpn, esor in lineas_requeridas:
            q_autr = (
                f"INSERT INTO AUTORIZADO (AUTR_CMPN_CODIGO, AUTR_ESOR_CODIGO, AUTR_CODIGO, AUTR_NOMBRE) "
                f"VALUES ('{cmpn}', '{esor}', '{nuevo_login}', '{nombre_completo}')"
            )
            ejecutar_actualizacion_db(q_autr, conn)
        
        # 3. INSERT autorizado_serie (copia masiva)
        q_serie = (
            f"INSERT INTO autorizado_serie (AUSR_CMPN_CODIGO, AUSR_ESOR_CODIGO, AUSR_SRDC_CODIGO, AUSR_DCMT_CODIGO, AUSR_AUTR_CODIGO, AUTR_NOMBRE) "
            f"SELECT a.AUSR_CMPN_CODIGO, a.AUSR_ESOR_CODIGO, a.AUSR_SRDC_CODIGO, a.AUSR_DCMT_CODIGO, '{nuevo_login}', '{nombre_completo}' "
            f"FROM autorizado_serie a WHERE a.AUSR_AUTR_CODIGO = '{usuario_ref}' "
            f"AND NOT EXISTS (SELECT 1 FROM autorizado_serie b WHERE b.AUSR_CMPN_CODIGO = a.AUSR_CMPN_CODIGO "
            f"AND b.AUSR_ESOR_CODIGO = a.AUSR_ESOR_CODIGO "
            f"AND b.AUSR_SRDC_CODIGO = a.AUSR_SRDC_CODIGO "
            f"AND b.AUSR_DCMT_CODIGO = a.AUSR_DCMT_CODIGO "
            f"AND b.AUSR_AUTR_CODIGO = '{nuevo_login}')"
        )
        ejecutar_actualizacion_db(q_serie, conn)
        
        return {
            "exito": True,
            "login": nuevo_login,
            "mensaje": f"Cédula {cedula} - Usuario {nuevo_login}: Creado exitosamente. Contraseña: {cedula}"
        }

def procesar_masiva(datos_extraidos, conn, page, frame, tipo_solicitud, id_ticket=None):
    """
    Procesa una solicitud masiva:
    1. Descarga el Excel adjunto del ticket
    2. Lee las filas (desde fila 13: B=Cédula, C=Nombre, F=Correo, G=Usuario ref)
    3. Aplica las validaciones de creación o activación por cada fila
    4. Publica un resumen consolidado en el ticket y lo finaliza
    """
    print("\n========================================")
    print("  PROCESAMIENTO MASIVO INICIADO")
    print("========================================")
    
    # 1. Descargar el Excel adjunto
    ruta_excel = descargar_adjunto_excel(page, frame)
    if not ruta_excel or not os.path.exists(ruta_excel):
        print("Error: No se pudo obtener el archivo Excel adjunto.")
        # Publicar nota de error y rechazar
        msg_error = ("Buen día.\nNo fue posible descargar o localizar el archivo Excel adjunto al ticket.\n"
                      "Por favor, verifique que el archivo esté adjunto correctamente y genere un nuevo requerimiento.")
        try:
            textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
            loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
            loc.wait_for(state="visible", timeout=10000)
            loc.fill(msg_error)
            btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
            btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
            btn_loc.wait_for(state="visible", timeout=5000)
            btn_loc.click()
            page.wait_for_timeout(2000)
        except Exception as e:
            print(f"Aviso: Error al publicar nota de fallo en descarga: {e}")
        return
    
    # 2. Leer datos del Excel
    usuarios = leer_excel_masivo(ruta_excel)
    if not usuarios:
        print("Error: El Excel no contiene datos de usuarios.")
        msg_vacio = ("Buen día.\nEl archivo Excel adjunto no contiene datos de usuarios a partir de la fila 13.\n"
                      "Por favor, verifique el formato del archivo y genere un nuevo requerimiento.")
        try:
            textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
            loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
            loc.wait_for(state="visible", timeout=10000)
            loc.fill(msg_vacio)
            btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
            btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
            btn_loc.wait_for(state="visible", timeout=5000)
            btn_loc.click()
            page.wait_for_timeout(2000)
        except Exception as e:
            print(f"Aviso: Error al publicar nota de Excel vacío: {e}")
        return
    
    # 3. Procesar cada fila según el tipo de solicitud
    resultados = []
    exitosos = 0
    fallidos = 0
    
    es_activacion = tipo_solicitud == "Activación/desbloqueo"
    es_creacion = "Creación" in tipo_solicitud or "Creacion" in tipo_solicitud
    
    for idx, usuario in enumerate(usuarios, start=1):
        print(f"\n--- Procesando usuario {idx}/{len(usuarios)} ---")
        cedula = usuario["cedula"]
        nombre = usuario["nombre"]
        correo = usuario["correo"]
        usuario_ref = usuario["usuario_referencia"]
        
        if es_activacion:
            resultado = validar_fila_activacion(cedula, usuario_ref, conn)
        elif es_creacion:
            resultado = validar_fila_creacion(cedula, nombre, correo, usuario_ref, conn)
        else:
            # Si el tipo no es claro, intentar activación primero
            print(f"  Tipo de solicitud no reconocido: '{tipo_solicitud}'. Intentando activación...")
            resultado = validar_fila_activacion(cedula, usuario_ref, conn)
        
        resultados.append(resultado)
        if resultado["exito"]:
            exitosos += 1
        else:
            fallidos += 1
    
    # 4. Construir resumen consolidado
    print("\n========================================")
    print("  RESUMEN DEL PROCESAMIENTO MASIVO")
    print(f"  Total: {len(usuarios)} | Exitosos: {exitosos} | Fallidos: {fallidos}")
    print("========================================")
    
    tipo_texto = "ACTIVACIÓN" if es_activacion else "CREACIÓN"
    resumen = f"PROCESAMIENTO MASIVO DE {tipo_texto}\n\n"
    resumen += "DETALLE POR USUARIO:\n"
    resumen += "-" * 50 + "\n"
    
    for idx, resultado in enumerate(resultados, start=1):
        resumen += f"{idx}. {resultado['mensaje']}\n"
    
    resumen += "-" * 50 + "\n"
    
    if es_activacion and exitosos > 0:
        resumen += "\nINSTRUCCIONES PARA LOS USUARIOS ACTIVADOS:\n"
        resumen += "Digita en la página inicial de Datasoft http://172.17.21.14/Datasoft/login.php los campos: Usuario, "
        resumen += "compañía, sin darle Clave y código de seguridad y luego das clic en Restablecer Contraseña.\n"
        resumen += "IMPORTANTE: Se debe ingresar al sistema antes de 24 horas para Activar el Usuario.\n"
        resumen += "Tener en cuenta al cambiar la contraseña:\n"
        resumen += "• Longitud mínima 8 caracteres\n"
        resumen += "• Longitud máxima 12 caracteres\n"
        resumen += "• Debe contener como mínimo (2) dos letras\n"
        resumen += "• Debe contener como mínimo (2) dos números\n"
        resumen += "• Debe contener como mínimo (2) dos caracteres especiales de la siguiente lista: $ - _ =\n"
        resumen += "• No se pueden repetir contraseñas anteriores.\n"
        resumen += "• No usar el asterisco * (asterisco)\n"
    
    if es_creacion and exitosos > 0:
        resumen += "\nINSTRUCCIONES PARA LOS USUARIOS CREADOS:\n"
        resumen += "POR FAVOR LEER MUY DESPACIO Y SEGUIR EL PASO A PASO CON LAS INSTRUCCIONES QUE SE DETALLAN A CONTINUACIÓN\n\n"
        resumen += "Digita en la página inicial de Datasoft http://172.17.21.14/Datasoft/login.php los campos: Usuario, "
        resumen += "compañía, sin darle Clave y código de seguridad y luego das clic en Restablecer Contraseña.\n"
        resumen += "IMPORTANTE: Se debe ingresar al sistema antes de 24 horas para Activar el Usuario.\n"
        resumen += "Tener en cuenta al cambiar la contraseña:\n"
        resumen += "• Longitud mínima 8 caracteres\n"
        resumen += "• Longitud máxima 12 caracteres\n"
        resumen += "• Debe contener como mínimo (2) dos letras\n"
        resumen += "• Debe contener como mínimo (2) dos números\n"
        resumen += "• Debe contener como mínimo (2) dos caracteres especiales de la siguiente lista: $ - _ =\n"
        resumen += "• No se pueden repetir contraseñas anteriores.\n"
        resumen += "• No usar el asterisco * (asterisco)\n"
    
    resumen += "\nSaludos,"
    
    # 5. Publicar resumen en el ticket de Helix
    try:
        textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
        loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
        loc.wait_for(state="visible", timeout=10000)
        loc.fill(resumen)
        
        # Clic en "Publicación"
        btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
        btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
        btn_loc.wait_for(state="visible", timeout=5000)
        btn_loc.click()
        print("Resumen masivo publicado en el ticket.")
        
        # Asignarme a mí
        page.wait_for_timeout(2000)
        btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
        btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
        btn_asignar_loc.wait_for(state="visible", timeout=10000)
        btn_asignar_loc.click()
        print("Ticket asignado correctamente.")
        
        # Editar → Cambiar estado
        page.wait_for_timeout(2000)
        btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
        btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
        btn_editar_loc.wait_for(state="visible", timeout=10000)
        btn_editar_loc.click()
        
        page.wait_for_timeout(1000)
        estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
        estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
        estado_loc.wait_for(state="visible", timeout=10000)
        estado_loc.click()
        
        # Si hubo fallos, marcar como Finalizado igualmente (el detalle queda en la nota)
        page.wait_for_timeout(1000)
        finalizado_sel = (
            'button.rx-select__option:has-text("Finalizado"), '
            'button[role="option"]:has-text("Finalizado"), '
            'button.rx-select__option:has-text("Completed"), '
            'button[role="option"]:has-text("Completed")'
        )
        finalizado_loc = frame.locator(finalizado_sel).first if frame else page.locator(finalizado_sel).first
        finalizado_loc.wait_for(state="visible", timeout=10000)
        finalizado_loc.click()
        
        page.wait_for_timeout(1000)
        guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
        guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
        guardar_loc.wait_for(state="visible", timeout=5000)
        guardar_loc.click()
        print("Ticket masivo finalizado y guardado exitosamente.")
    except Exception as e:
        print(f"Aviso: Error al publicar resumen o finalizar ticket masivo: {e}")
    
    # 6. Generar informe en GLPI
    if page:
        titulo_glpi = f"{tipo_texto} USUARIO BANCO - MASIVO ({exitosos}/{len(usuarios)} exitosos)"
        generar_informe_glpi(page, datos_extraidos, None, None, "Finalizado", titulo_glpi, id_ticket=id_ticket)
    
    # 7. Limpiar archivo temporal
    try:
        if ruta_excel and os.path.exists(ruta_excel) and tempfile.gettempdir() in ruta_excel:
            os.remove(ruta_excel)
            print("Archivo temporal eliminado.")
    except:
        pass
    
    print("\n========================================")
    print("  PROCESAMIENTO MASIVO COMPLETADO")
    print("========================================")

def validar_datos_condicionales(datos_extraidos, conn, page=None, frame=None, id_ticket=None):
    print("\n--- Validando en Base de Datos vía Terminal ---")

    info_servicio = datos_extraidos.get("Información del servicio", "").lower()
    tipo_solicitud = datos_extraidos.get("Tipo solicitud", "")
    usuario_datasoft = datos_extraidos.get("Usuario de Datasoft", "")
    cedula = datos_extraidos.get("No. Cédula", "")

    if "masiva" in info_servicio:
        print("Solicitud masiva detectada. Procesando desde Excel adjunto...")
        procesar_masiva(datos_extraidos, conn, page, frame, tipo_solicitud, id_ticket=id_ticket)
        return

    try:
        if tipo_solicitud == "Activación/desbloqueo":
            print(f"Validando existencia de '{usuario_datasoft}' o '{cedula}'...")
            existe = False
            login_final = usuario_datasoft
            
            if usuario_datasoft:
                res = ejecutar_consulta_db(f"SELECT COUNT(*) FROM a_usuario WHERE LOGIN = '{usuario_datasoft}'", conn)
                if res and res.strip() != "0":
                    existe = True
            
            if not existe and cedula:
                res_login = ejecutar_consulta_db(f"SELECT LOGIN FROM a_usuario WHERE NO_IDENTIFICACION = '{cedula}'", conn)
                if res_login and res_login.strip() != "0":
                    existe = True
                    login_final = res_login
                    print(f"Se comprobó existencia por cédula. Usando LOGIN de la base: '{login_final}'")
                    
            if existe:
                print("existe")
                if login_final:
                    print(f"Actualizando usuario {login_final}...")
                    update_query = f"UPDATE a_usuario SET CONTRASENA = NO_IDENTIFICACION, ESTADO = 'A', CAMBIO_CONTRASENA = 'S' WHERE LOGIN = '{login_final}'"
                    exito = ejecutar_actualizacion_db(update_query, conn)
                    if exito:
                        print("Actualización exitosa. Ingresando nota al ticket...")
                        mensaje = f"""POR FAVOR LEER MUY DESPACIO Y SEGUIR EL  PASO A PASO CON LAS INSTRUCCIONES QUE SE DETALLAN A CONTINUACIÓN

Buen día.

Tu usuario fue activado;
Usuario: {login_final}
Contraseña: {cedula}

Debes realizar los siguientes pasos para restablecer la contraseña o cambiarla:

Digita en la página inicial de Datasoft http://172.17.21.14/Datasoft/login.php los campos: Usuario, 
compañía, sin darle Clave y código de seguridad y luego das clic en Restablecer Contraseña.
IMPORTANTE: Se debe ingresar al sistema antes de 24 horas para Activar el Usuario y evitar que se vuelva a 
inactivar. Tener en cuenta al cambiar la contraseña:
• Longitud mínima 8 caracteres
• Longitud máxima 12 caracteres
• Debe contener como mínimo (2) dos letras
• Debe contener como mínimo (2) dos números
• Debe contener como mínimo (2) dos caracteres especiales de la siguiente lista: $ - _ =
• No se pueden repetir contraseñas anteriores.
• No usar el asterisco * (asterisco)

Saludos,"""
                        try:
                            # Intentar escribir la nota en el textarea dentro del frame o la página
                            textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
                            loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
                            loc.wait_for(state="visible", timeout=10000)
                            loc.fill(mensaje)
                            
                            # Clic en el botón "Publicación"
                            btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
                            btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
                            btn_loc.wait_for(state="visible", timeout=5000)
                            btn_loc.click()
                            print("Nota ingresada y publicada correctamente en el ticket.")
                            
                            # Clic en "Asignarme a mí"
                            print("Asignando el ticket a mí mismo...")
                            page.wait_for_timeout(2000) # Esperar a que consolide la nota
                            btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
                            btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
                            btn_asignar_loc.wait_for(state="visible", timeout=10000)
                            btn_asignar_loc.click()
                            print("Ticket asignado correctamente.")
                            
                            # 1. Clic en "Editar"
                            print("Cambiando el estado a Finalizado...")
                            page.wait_for_timeout(2000) # Esperar a que consolide la asignación
                            btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
                            btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
                            btn_editar_loc.wait_for(state="visible", timeout=5000)
                            btn_editar_loc.click()
                            
                            # 2. Clic en select "Estado"
                            page.wait_for_timeout(1000)
                            estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
                            estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
                            estado_loc.wait_for(state="visible", timeout=10000)
                            estado_loc.click()
                            
                            # 3. Seleccionar "Finalizado"
                            page.wait_for_timeout(2000)
                            finalizado_sel = (
                                'button.rx-select__option:has-text("Finalizado"), '
                                'button[role="option"]:has-text("Finalizado"), '
                                'button.rx-select__option:has-text("Completed"), '
                                'button[role="option"]:has-text("Completed")'
                            )
                            finalizado_loc = frame.locator(finalizado_sel).first if frame else page.locator(finalizado_sel).first
                            finalizado_loc.wait_for(state="visible", timeout=10000)
                            finalizado_loc.click()
                            
                            # 4. Clic en "Guardar ticket"
                            page.wait_for_timeout(1000)
                            guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
                            guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
                            guardar_loc.wait_for(state="visible", timeout=5000)
                            guardar_loc.click()
                            print("Ticket finalizado y guardado exitosamente.")
                        except Exception as e:
                            print(f"Aviso: No se pudo ingresar o publicar la nota en el textarea. Detalle: {e}")
                    else:
                        print("Fallo la actualización.")
                else:
                    print("Aviso: No se pudo actualizar porque falta el Usuario de Datasoft.")
            else:
                print("no existe")
                print("Ingresando nota de rechazo al ticket...")
                mensaje_rechazo = """Buen dia.
Tras validar el ticket, confirmamos que la cuenta solicitada no existe. 
Por este motivo, procedemos con el rechazo de la solicitud. 
Por favor, genere un nuevo requerimiento para la creacion del usuario."""
                try:
                    # 1. Escribir nota
                    textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
                    loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
                    loc.wait_for(state="visible", timeout=10000)
                    loc.fill(mensaje_rechazo)
                    
                    # 2. Clic en "Publicación"
                    btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
                    btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
                    btn_loc.wait_for(state="visible", timeout=5000)
                    btn_loc.click()
                    print("Nota de rechazo ingresada y publicada.")
                    
                    # Clic en "Asignarme a mí"
                    print("Asignando el ticket a mí mismo...")
                    page.wait_for_timeout(2000)
                    btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
                    btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
                    btn_asignar_loc.wait_for(state="visible", timeout=10000)
                    btn_asignar_loc.click()
                    print("Ticket asignado correctamente.")
                    
                    # 3. Clic en "Editar"
                    print("Cambiando el estado a Rechazado...")
                    page.wait_for_timeout(2000)
                    btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
                    btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
                    btn_editar_loc.wait_for(state="visible", timeout=10000)
                    btn_editar_loc.click()
                    
                    # 4. Clic en select "Estado"
                    page.wait_for_timeout(1000)
                    estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
                    estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
                    estado_loc.wait_for(state="visible", timeout=10000)
                    estado_loc.click()
                    
                    # 5. Seleccionar "Rechazado"
                    page.wait_for_timeout(1000)
                    rechazado_sel = (
                        'button.rx-select__option:has-text("Rechazado"), '
                        'button[role="option"]:has-text("Rechazado"), '
                        'button.rx-select__option:has-text("Rejected"), '
                        'button[role="option"]:has-text("Rejected")'
                    )
                    rechazado_loc = frame.locator(rechazado_sel).first if frame else page.locator(rechazado_sel).first
                    rechazado_loc.wait_for(state="visible", timeout=5000)
                    rechazado_loc.click()
                    
                    # 6. Clic en "Guardar ticket"
                    page.wait_for_timeout(1000)
                    guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
                    guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
                    guardar_loc.wait_for(state="visible", timeout=5000)
                    guardar_loc.click()
                    print("Ticket rechazado y guardado exitosamente.")
                except Exception as e:
                    print(f"Aviso: No se pudo ingresar la nota o registrar el rechazo. Detalle: {e}")
                    
            # Generar el registro en GLPI al finalizar independientemente de si existía o no
            if page:
                login_a_enviar = login_final if existe else None
                pass_a_enviar = cedula if existe else None
                estado_h = "Finalizado" if existe else "Rechazado"
                generar_informe_glpi(page, datos_extraidos, login_a_enviar, pass_a_enviar, estado_h, id_ticket=id_ticket)
                
        elif "Creación" in tipo_solicitud or "Creacion" in tipo_solicitud:
            print(f"Validando creacion de '{cedula}'...")
            nombre_completo  = datos_extraidos.get("Nombre Completo", "")
            email_usuario    = datos_extraidos.get("Correo", "")
            usuario_ref      = datos_extraidos.get("Usuario de referencia DATASOFT", "")
            linea_negocio    = datos_extraidos.get("Línea de Negocio", "")

            # ── 1. Validar que el usuario de referencia exista ──────────────────
            ref_fila = ejecutar_consulta_fila_db(
                f"SELECT LOGIN, CMPN_CODIGO, ESOR_CODIGO, NOMBRE_USUARIO FROM a_usuario WHERE LOGIN = '{usuario_ref}'",
                conn
            )
            if not ref_fila:
                print("Usuario de referencia no existe. Rechazando...")
                msg = ("Buen dia.\nTras validar el ticket, confirmamos que la cuenta solicitada "
                       "del Usuario de referencia no existe.\n"
                       "Por este motivo, procedemos con el rechazo de la solicitud.\n"
                       "Por favor, genere un nuevo requerimiento con un Usuario de referencia Correcto.")
                try:
                    textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
                    loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
                    loc.wait_for(state="visible", timeout=10000)
                    loc.fill(msg)
                    btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
                    btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
                    btn_loc.wait_for(state="visible", timeout=5000)
                    btn_loc.click()
                    page.wait_for_timeout(2000)
                    btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
                    btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
                    btn_asignar_loc.wait_for(state="visible", timeout=10000)
                    btn_asignar_loc.click()
                    page.wait_for_timeout(2000)
                    btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
                    btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
                    btn_editar_loc.wait_for(state="visible", timeout=10000)
                    btn_editar_loc.click()
                    page.wait_for_timeout(1000)
                    estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
                    estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
                    estado_loc.wait_for(state="visible", timeout=10000)
                    estado_loc.click()
                    page.wait_for_timeout(1000)
                    rechazado_sel = (
                        'button.rx-select__option:has-text("Rechazado"), '
                        'button[role="option"]:has-text("Rechazado"), '
                        'button.rx-select__option:has-text("Rejected"), '
                        'button[role="option"]:has-text("Rejected")'
                    )
                    rechazado_loc = frame.locator(rechazado_sel).first if frame else page.locator(rechazado_sel).first
                    rechazado_loc.wait_for(state="visible", timeout=5000)
                    rechazado_loc.click()
                    page.wait_for_timeout(1000)
                    guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
                    guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
                    guardar_loc.wait_for(state="visible", timeout=5000)
                    guardar_loc.click()
                    print("Ticket rechazado por usuario de referencia inexistente.")
                except Exception as e:
                    print(f"Aviso: Error al rechazar ticket por ref inexistente: {e}")
                if page:
                    generar_informe_glpi(page, datos_extraidos, None, None, "Rechazado", "CREACION USUARIO BANCO", id_ticket=id_ticket)
                return

            ref_cmpn = ref_fila.get("CMPN_CODIGO", "BANCOLOMBI")
            ref_esor = ref_fila.get("ESOR_CODIGO", "PIC")

            # ── 2. Verificar si usuario ya existe por cédula ─────────────────────
            login_existente = ejecutar_consulta_db(
                f"SELECT LOGIN FROM a_usuario WHERE NO_IDENTIFICACION = '{cedula}'", conn
            )
            usuario_ya_existe = login_existente and login_existente.strip() not in ("0", "")

            # Líneas de negocio que corresponden según Helix
            lineas_requeridas = lineas_desde_helix(linea_negocio)

            if usuario_ya_existe:
                login_existente = login_existente.strip()
                print(f"Usuario ya existe: {login_existente}. Verificando autorizaciones...")

                # Verificar qué líneas ya tiene en AUTORIZADO
                lineas_faltantes = []
                for cmpn, esor in lineas_requeridas:
                    res_autr = ejecutar_consulta_db(
                        f"SELECT COUNT(*) FROM AUTORIZADO WHERE AUTR_CODIGO = '{login_existente}' AND AUTR_CMPN_CODIGO = '{cmpn}'",
                        conn
                    )
                    if res_autr and res_autr.strip() == "0":
                        lineas_faltantes.append((cmpn, esor))

                if not lineas_faltantes:
                    # ── Escenario B1: ya tiene todo → Rechazar ─────────────────
                    print("El usuario ya tiene todas las autorizaciones solicitadas. Rechazando...")
                    msg = ("Buen dia.\nTras validar el ticket, confirmamos que la cuenta solicitada ya existe "
                           "y cuenta con la Línea de Negocio indicada.\n"
                           "Por este motivo, procedemos con el rechazo de la solicitud.")
                    try:
                        textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
                        loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
                        loc.wait_for(state="visible", timeout=10000)
                        loc.fill(msg)
                        btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
                        btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
                        btn_loc.wait_for(state="visible", timeout=5000)
                        btn_loc.click()
                        page.wait_for_timeout(2000)
                        btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
                        btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
                        btn_asignar_loc.wait_for(state="visible", timeout=10000)
                        btn_asignar_loc.click()
                        page.wait_for_timeout(2000)
                        btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
                        btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
                        btn_editar_loc.wait_for(state="visible", timeout=10000)
                        btn_editar_loc.click()
                        page.wait_for_timeout(1000)
                        estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
                        estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
                        estado_loc.wait_for(state="visible", timeout=10000)
                        estado_loc.click()
                        page.wait_for_timeout(1000)
                        rechazado_sel = (
                            'button.rx-select__option:has-text("Rechazado"), '
                            'button[role="option"]:has-text("Rechazado"), '
                            'button.rx-select__option:has-text("Rejected"), '
                            'button[role="option"]:has-text("Rejected")'
                        )
                        rechazado_loc = frame.locator(rechazado_sel).first if frame else page.locator(rechazado_sel).first
                        rechazado_loc.wait_for(state="visible", timeout=5000)
                        rechazado_loc.click()
                        page.wait_for_timeout(1000)
                        guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
                        guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
                        guardar_loc.wait_for(state="visible", timeout=5000)
                        guardar_loc.click()
                        print("Ticket rechazado: usuario ya existe con esa línea de negocio.")
                    except Exception as e:
                        print(f"Aviso: Error al rechazar ticket (Escenario B1): {e}")
                    if page:
                        generar_informe_glpi(page, datos_extraidos, None, None, "Rechazado", "CREACION USUARIO BANCO", id_ticket=id_ticket)
                    return

                else:
                    # ── Escenario B2: falta alguna línea → Añadir AUTORIZADO + autorizado_serie ──
                    print(f"Añadiendo {len(lineas_faltantes)} autorización(es) faltante(s)...")
                    nombre_ref = ref_fila.get("NOMBRE_USUARIO", nombre_completo)
                    for cmpn, esor in lineas_faltantes:
                        q_autr = (f"INSERT INTO AUTORIZADO (AUTR_CMPN_CODIGO, AUTR_ESOR_CODIGO, AUTR_CODIGO, AUTR_NOMBRE) "
                                  f"VALUES ('{cmpn}', '{esor}', '{login_existente}', '{nombre_completo}')")
                        ejecutar_actualizacion_db(q_autr, conn)
                        q_serie = (f"INSERT INTO autorizado_serie (AUSR_CMPN_CODIGO, AUSR_ESOR_CODIGO, AUSR_SRDC_CODIGO, AUSR_DCMT_CODIGO, AUSR_AUTR_CODIGO, AUTR_NOMBRE) "
                                   f"SELECT a.AUSR_CMPN_CODIGO, a.AUSR_ESOR_CODIGO, a.AUSR_SRDC_CODIGO, a.AUSR_DCMT_CODIGO, '{login_existente}', '{nombre_completo}' "
                                   f"FROM autorizado_serie WHERE AUSR_AUTR_CODIGO = '{usuario_ref}'"
                                   f"AND NOT EXISTS (SELECT 1 FROM autorizado_serie b WHERE b.AUSR_CMPN_CODIGO = a.AUSR_CMPN_CODIGO "
                                   f"AND b.AUSR_ESOR_CODIGO = a.AUSR_ESOR_CODIGO "
                                   f"AND b.AUSR_SRDC_CODIGO = a.AUSR_SRDC_CODIGO "
                                   f"AND b.AUSR_DCMT_CODIGO = a.AUSR_DCMT_CODIGO "
                                   f"AND b.AUSR_AUTR_CODIGO = '{login_existente}')")
                        ejecutar_actualizacion_db(q_serie, conn)
                    print("Autorizaciones añadidas exitosamente.")

                    msg_exito = (f"Buen dia.\nTras validar el ticket, confirmamos que la cuenta {login_existente} "
                                 f"ya existía y se le ha añadido la Línea de Negocio solicitada ({linea_negocio}).\n"
                                 f"El usuario puede iniciar sesión con su contraseña actual.")
                    try:
                        textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
                        loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
                        loc.wait_for(state="visible", timeout=10000)
                        loc.fill(msg_exito)
                        btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
                        btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
                        btn_loc.wait_for(state="visible", timeout=5000)
                        btn_loc.click()
                        page.wait_for_timeout(2000)
                        btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
                        btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
                        btn_asignar_loc.wait_for(state="visible", timeout=10000)
                        btn_asignar_loc.click()
                        page.wait_for_timeout(2000)
                        btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
                        btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
                        btn_editar_loc.wait_for(state="visible", timeout=10000)
                        btn_editar_loc.click()
                        page.wait_for_timeout(1000)
                        estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
                        estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
                        estado_loc.wait_for(state="visible", timeout=10000)
                        estado_loc.click()
                        page.wait_for_timeout(1000)
                        finalizado_sel = (
                            'button.rx-select__option:has-text("Finalizado"), '
                            'button[role="option"]:has-text("Finalizado"), '
                            'button.rx-select__option:has-text("Completed"), '
                            'button[role="option"]:has-text("Completed")'
                        )
                        finalizado_loc = frame.locator(finalizado_sel).first if frame else page.locator(finalizado_sel).first
                        finalizado_loc.wait_for(state="visible", timeout=5000)
                        finalizado_loc.click()
                        page.wait_for_timeout(1000)
                        guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
                        guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
                        guardar_loc.wait_for(state="visible", timeout=5000)
                        guardar_loc.click()
                        print("Ticket finalizado: autorización añadida.")
                    except Exception as e:
                        print(f"Aviso: Error al finalizar ticket (Escenario B2): {e}")
                    if page:
                        generar_informe_glpi(page, datos_extraidos, login_existente, cedula, "Finalizado", "CREACION USUARIO BANCO", id_ticket=id_ticket)
                    return

            else:
                # ── Escenario A: Usuario no existe → Creación Completa ─────────
                print("Usuario no existe. Creando...")
                nuevo_login = generar_login(nombre_completo, conn)
                print(f"LOGIN generado: {nuevo_login}")

                # 1. INSERT a_usuario
                q_usuario = (
                    f"INSERT INTO a_usuario (ID_USUARIO, LOGIN, CONTRASENA, ID_GRUPO_USUARIO, NOMBRE_USUARIO, "
                    f"ESTADO, LOGIN_USUARIO, FECHA_ULTIMA_ACT, CMPN_CODIGO, ESOR_CODIGO, E_MAIL, "
                    f"TIPO_USUARIO, CAMBIO_CONTRASENA, FECHA_CAMBIO_CONTRASENA, NO_INTENTOS, "
                    f"RESTABLECE_CONTRASENA, NO_IDENTIFICACION, FECHA_CREACION) "
                    f"VALUES (seq_a_usuario.nextval, '{nuevo_login}', '{cedula}', 462, "
                    f"'{nombre_completo}', 'A', 'SPINE', SYSDATE, '{ref_cmpn}', '{ref_esor}', "
                    f"'{email_usuario}', 'E', 'N', SYSDATE, 0, 'S', '{cedula}', SYSDATE)"
                )
                exito_usuario = ejecutar_actualizacion_db(q_usuario, conn)
                if not exito_usuario:
                    print("Error: No se pudo crear el usuario en a_usuario. Abortando.")
                    return
                print("Usuario creado en a_usuario exitosamente.")

                # 2. INSERT AUTORIZADO
                for cmpn, esor in lineas_requeridas:
                    q_autr = (f"INSERT INTO AUTORIZADO (AUTR_CMPN_CODIGO, AUTR_ESOR_CODIGO, AUTR_CODIGO, AUTR_NOMBRE) "
                              f"VALUES ('{cmpn}', '{esor}', '{nuevo_login}', '{nombre_completo}')")
                    ejecutar_actualizacion_db(q_autr, conn)
                print(f"{len(lineas_requeridas)} registro(s) insertado(s) en AUTORIZADO.")

                # 3. INSERT autorizado_serie (copia masiva)
                q_serie = (
                    f"INSERT INTO autorizado_serie (AUSR_CMPN_CODIGO, AUSR_ESOR_CODIGO, AUSR_SRDC_CODIGO, AUSR_DCMT_CODIGO, AUSR_AUTR_CODIGO, AUTR_NOMBRE) "
                    f"SELECT a.AUSR_CMPN_CODIGO, a.AUSR_ESOR_CODIGO, a.AUSR_SRDC_CODIGO, a.AUSR_DCMT_CODIGO, '{nuevo_login}', '{nombre_completo}' "
                    f"FROM autorizado_serie a WHERE AUSR_AUTR_CODIGO = '{usuario_ref}'"
                    f"AND NOT EXISTS (SELECT 1 FROM autorizado_serie b WHERE b.AUSR_CMPN_CODIGO = a.AUSR_CMPN_CODIGO "
                    f"AND b.AUSR_ESOR_CODIGO = a.AUSR_ESOR_CODIGO "
                    f"AND b.AUSR_SRDC_CODIGO = a.AUSR_SRDC_CODIGO "
                    f"AND b.AUSR_DCMT_CODIGO = a.AUSR_DCMT_CODIGO "
                    f"AND b.AUSR_AUTR_CODIGO = '{nuevo_login}')"
                )
                exito_serie = ejecutar_actualizacion_db(q_serie, conn)
                if exito_serie:
                    print("Permisos de autorizado_serie copiados exitosamente.")

                # Nota en Helix + finalizar
                mensaje_creacion = (
                    f"POR FAVOR LEER MUY DESPACIO Y SEGUIR EL PASO A PASO CON LAS INSTRUCCIONES QUE SE DETALLAN A CONTINUACIÓN\n\n"
                    f"Buen día.\n\n"
                    f"Tu usuario ha sido creado exitosamente;\n"
                    f"Usuario: {nuevo_login}\n"
                    f"Contraseña: {cedula}\n\n"
                    f"Debes realizar los siguientes pasos para restablecer la contraseña o cambiarla:\n\n"
                    f"Digita en la página inicial de Datasoft http://172.17.21.14/Datasoft/login.php los campos: Usuario, "
                    f"compañía, sin darle Clave y código de seguridad y luego das clic en Restablecer Contraseña.\n"
                    f"IMPORTANTE: Se debe ingresar al sistema antes de 24 horas para Activar el Usuario y evitar que se vuelva a "
                    f"inactivar. Tener en cuenta al cambiar la contraseña:\n"
                    f"• Longitud mínima 8 caracteres\n"
                    f"• Longitud máxima 12 caracteres\n"
                    f"• Debe contener como mínimo (2) dos letras\n"
                    f"• Debe contener como mínimo (2) dos números\n"
                    f"• Debe contener como mínimo (2) dos caracteres especiales de la siguiente lista: $ - _ =\n"
                    f"• No se pueden repetir contraseñas anteriores.\n"
                    f"• No usar el asterisco * (asterisco)\n\n"
                    f"Saludos,"
                )
                try:
                    textarea_selector = 'textarea[data-testid="304247080"], textarea[name="ar304247080"]'
                    loc = frame.locator(textarea_selector).first if frame else page.locator(textarea_selector).first
                    loc.wait_for(state="visible", timeout=10000)
                    loc.fill(mensaje_creacion)
                    btn_selector = 'button[name="ar304268430"], button[id="304268430"], button[title="Publicación"]'
                    btn_loc = frame.locator(btn_selector).first if frame else page.locator(btn_selector).first
                    btn_loc.wait_for(state="visible", timeout=5000)
                    btn_loc.click()
                    print("Nota de creación publicada en Helix.")
                    page.wait_for_timeout(2000)
                    btn_asignar_sel = 'button[name="ar304421551"], button[id="304421551"], button[title="Asignarme a mí"]'
                    btn_asignar_loc = frame.locator(btn_asignar_sel).first if frame else page.locator(btn_asignar_sel).first
                    btn_asignar_loc.wait_for(state="visible", timeout=10000)
                    btn_asignar_loc.click()
                    page.wait_for_timeout(2000)
                    btn_editar_sel = 'button[name="ar304420591"], button[title="Editar"]'
                    btn_editar_loc = frame.locator(btn_editar_sel).first if frame else page.locator(btn_editar_sel).first
                    btn_editar_loc.wait_for(state="visible", timeout=10000)
                    btn_editar_loc.click()
                    page.wait_for_timeout(1000)
                    estado_sel = 'button[name="ar7"], button[aria-label="Estado"]'
                    estado_loc = frame.locator(estado_sel).first if frame else page.locator(estado_sel).first
                    estado_loc.wait_for(state="visible", timeout=10000)
                    estado_loc.click()
                    page.wait_for_timeout(1000)
                    finalizado_sel = (
                        'button.rx-select__option:has-text("Finalizado"), '
                        'button[role="option"]:has-text("Finalizado"), '
                        'button.rx-select__option:has-text("Completed"), '
                        'button[role="option"]:has-text("Completed")'
                    )
                    finalizado_loc = frame.locator(finalizado_sel).first if frame else page.locator(finalizado_sel).first
                    finalizado_loc.wait_for(state="visible", timeout=5000)
                    finalizado_loc.click()
                    page.wait_for_timeout(1000)
                    guardar_sel = 'button[name="ar304440891"], button[title="Guardar ticket"]'
                    guardar_loc = frame.locator(guardar_sel).first if frame else page.locator(guardar_sel).first
                    guardar_loc.wait_for(state="visible", timeout=5000)
                    guardar_loc.click()
                    print("Ticket de creación finalizado y guardado.")
                except Exception as e:
                    print(f"Aviso: Error al finalizar ticket de creación: {e}")
                if page:
                    generar_informe_glpi(page, datos_extraidos, nuevo_login, cedula, "Finalizado", "CREACION USUARIO BANCO", id_ticket=id_ticket)
    except Exception as e:
        print(f"Error en la consulta: {e}")

def get_input_popup(prompt_text, is_password=False):
    root = tk.Tk()
    root.withdraw() # Ocultar la ventana principal
    root.attributes('-topmost', 1) # Asegurar que el popup salga al frente
    if is_password:
        result = simpledialog.askstring("Ingreso requerido", prompt_text, parent=root, show='*')
    else:
        result = simpledialog.askstring("Ingreso requerido", prompt_text, parent=root)
    root.destroy()
    return result

def validar_items(page, conn=None):
    print("=== Validando existencia de items en la tabla ===")
    try:
        # Esperamos a que aparezca la tabla vacía o la tabla con datos
        page.wait_for_selector(".tc__list-placeholder-text, .ngViewport", state="visible", timeout=30000)
        
        # Validamos cuál elemento cargó
        if page.locator(".tc__list-placeholder-text").is_visible():
            print("no hay items")
        elif page.locator(".ngViewport").is_visible():
            contar_items(page, conn)
        else:
            # Por si algo raro pasó con el DOM
            print("no se pudo determinar el estado de la tabla")
            
    except Exception as e:
        print(f"Aviso: Timeout al esperar la tabla. Detalle: {str(e)[:50]}")

def contar_items(page, conn=None):
    # Contar items iterando por fila
    count = 0
    while True:
        selector = f".ng-scope:nth-child({count + 1}) > .col2 .ngCellText"
        if page.locator(selector).count() > 0:
            count += 1
        else:
            break
    print(f"hay items: {count} item(s) encontrado(s)")

    # Obtener detalle de todos los ítems encontrados
    # IMPORTANTE: siempre hacemos clic en el ítem en posición 1, porque al procesar
    # y finalizar/rechazar un ticket, este desaparece de la lista y el siguiente
    # sube a ocupar su lugar. Usar un índice fijo (2, 3...) causaría un error.
    for i in range(count):
        print(f"Procesando ítem {i + 1} de {count} (siempre se toma el primero de la lista)...")
        obtener_detalle_item(page, item_index=1, conn=conn)
        
        # Si no es el último ítem, volver a la consola de tickets para poder abrir el siguiente
        if i < count - 1:
            print("Regresando a la consola de tickets...")
            page.go_back()
            # Esperar a que recargue la tabla original antes de buscar el siguiente
            page.wait_for_selector(".tc__list-placeholder-text, .ngViewport", state="visible", timeout=30000)
            page.wait_for_timeout(2000) # Pausa mínima para estabilizar la tabla


def obtener_detalle_item(page, item_index=1, conn=None):
    """
    Hace clic en el ítem de la lista, entra al iframe (#pwa-frame) 
    y extrae el valor específico de 'No. Cédula:'.
    """
    print(f"\n=== Extrayendo detalle del ítem #{item_index} ===")
    try:
        # 0. Extraer el ID del ticket (que empieza por WO) antes de entrar
        id_ticket_selector = f".ng-scope:nth-child({item_index}) > .col2 .ngCellText"
        id_ticket = "No encontrado"
        try:
            page.locator(id_ticket_selector).first.wait_for(state="visible", timeout=10000)
            id_ticket = page.locator(id_ticket_selector).first.inner_text().strip()
            print(f"ID del Ticket Helix detectado: {id_ticket}")
        except Exception as e:
            print(f"Aviso: No se pudo extraer el ID del ticket: {e}")

        # 1. Hacer clic en la fila para entrar al ítem
        fila_selector = f".ng-scope:nth-child({item_index}) > .col2 .ngCellText"
        page.locator(fila_selector).first.click()
        print("Clic realizado, entrando a la vista del ticket...")

        # 2. Entrar al iframe que contiene la vista progresiva de Helix
        # Hemos quitado las pausas largas! Dejamos que Playwright detecte el iframe rápido.
        frame = page.frame_locator("#pwa-frame")
        
        # 3. Esperar y ubicar el cuadro de texto exacto
        cuadro_datos = frame.locator("#ar1000000151_data")
        # El programa avanzará en el instante microscópico en que detecte que el texto existe
        cuadro_datos.first.wait_for(state="attached", timeout=20000)
        
        # Extraemos el texto completo (del atributo title que contiene la data estructurada)
        texto = cuadro_datos.first.get_attribute("title")
        if not texto:  # Respaldo por si acaso
            texto = cuadro_datos.first.inner_text()

        # 4. Aislar todos los campos solicitados
        campos_buscados = [
            "No. Cédula:",
            "Tipo solicitud:",
            "Usuario de Datasoft:",
            "Información del servicio:",
            "Usuario de referencia DATASOFT:",
            "Línea de Negocio:"
        ]
        
        datos_extraidos = {}
        
        for linea in texto.splitlines():
            for campo in campos_buscados:
                if campo in linea:
                    partes = linea.split(campo)
                    if len(partes) > 1:
                        # Guardamos el valor limpio sin los dos puntos
                        nombre_amigable = campo.replace(":", "").strip()
                        datos_extraidos[nombre_amigable] = partes[1].strip()
                        break

        # Extraer Nombre Completo
        try:
            nombre_loc = frame.locator("#ar301395400_data")
            if nombre_loc.count() == 0:
                nombre_loc = page.locator("#ar301395400_data")
            if nombre_loc.count() > 0:
                datos_extraidos["Nombre Completo"] = nombre_loc.first.inner_text().strip()
        except Exception as e:
            print(f"Aviso: No se pudo extraer Nombre Completo {e}")

        # Extraer Correo
        try:
            correo_loc = frame.locator("#ar1000000048_data")
            if correo_loc.count() == 0:
                correo_loc = page.locator("#ar1000000048_data")
            if correo_loc.count() > 0:
                datos_extraidos["Correo"] = correo_loc.first.inner_text().strip()
        except Exception as e:
            print(f"Aviso: No se pudo extraer Correo {e}")

        # 5. Imprimir el resultado sin caracteres especiales/emojis
        print("\n=================================")
        if datos_extraidos:
            print("DATOS EXTRAIDOS DEL ITEM:")
            # Se imprime sin acentos extraños en los prints para cuidar que la consola de Windows no falle
            for clave, valor in datos_extraidos.items():
                # Reemplazamos acentos en las llaves solo para la impresion segura en CMD
                clave_segura = clave.replace('é', 'e').replace('í', 'i')
                print(f"  {clave_segura}: {valor}")
                
            validar_datos_condicionales(datos_extraidos, conn, page, frame, id_ticket=id_ticket)
        else:
            print("Fallo: No se vio ningun texto de los solicitados.")
            print("Texto extraido para depurar:\n", texto)
        print("=================================\n")

    except Exception as e:
        print(f"Error al extraer los datos: {str(e)[:150]}")
    

@task
def login_smartit():
    print("--- Inicio de Sesión Smart IT ---")
    # Credenciales por defecto
    email = "spinerez@bancolombia.com.co"
    password = "D4taf1l3$M3d26#"

    # Configurar el navegador
    browser.configure(
        browser_engine="chromium",
        headless=False, # Necesitamos ver el navegador
        isolated=False,  # Modo Incógnito / Contexto limpio
    )
    
    print("Abriendo el navegador...")
    page = browser.goto("https://bancolombia-smartit.onbmc.com/smartit/app/#/ticket-console")
    
    # 1. Ingresar Correo
    print("Esperando y llenando campo de correo (las redirecciones SSO pueden tomar algo de tiempo)...")
    page.wait_for_selector("#i0116", state="visible", timeout=60000)
    page.locator("#i0116").fill(email)
    
    # Clic en "Siguiente"
    page.locator("#idSIButton9").click()
    
    # 2. Ingresar Contraseña
    print("Esperando y llenando contraseña...")
    page.wait_for_selector("#i0118", state="visible", timeout=30000)
    page.locator("#i0118").fill(password)
    
    # Clic en "Iniciar sesión" / Siguiente
    page.locator("#idSIButton9").click()
    
    # 3. Código del autenticador
    print("\nSe ha enviado la contraseña...")
    
    # Pedimos el código con popup mientras carga la página
    auth_code = get_input_popup("Por favor ingrese el código del autenticador que recibió:")
    
    if auth_code:
        print("Ingresando el código de autenticación...")
        try:
            page.wait_for_selector("#idTxtBx_SAOTCC_OTC", state="visible", timeout=30000)
            page.locator("#idTxtBx_SAOTCC_OTC").fill(auth_code)
            
            # Clic en "Confirmar" / "Verificar"
            page.locator("#idSubmit_SAOTCC_Continue").click()
        except Exception as e:
            print(f"Aviso: Ocurrió un error al intentar ingresar el código. Detalle: {str(e)[:50]}")
    
    # 4. Seleccionar cuenta si aparece el picker de cuentas de Microsoft (Falta verificar aveces aparece esta  raro)
    try:
        account_tile_selector = f"div[data-test-id='{email}']"
        page.wait_for_selector(account_tile_selector, state="visible", timeout=5000)
        print("Selector de cuenta detectado, haciendo click...")
        page.locator(account_tile_selector).click()
    except Exception:
        pass  # No apareció el selector de cuenta, se continúa normalmente
    
    # 5. Esperar a que cargue la consola de Smart IT
    print("Esperando a que cargue completamente la consola de Smart IT...")
    page.wait_for_timeout(10000) 
    
    # 5.5 Conectar a la base de datos Oracle
    conn = conectar_bd()
    
    # 6. Validar existencia de ítems
    validar_items(page, conn)
    
    # Cerrar conexion a BD al finalizar
    if conn:
        try:
            conn.close()
            print("Conexión a BD finalizada.")
        except Exception:
            pass
    
    # Tomar captura de pantalla
    os.makedirs("output/img", exist_ok=True)
    page.screenshot(path="output/img/smartit_login.png")
    print("Captura de pantalla guardada en 'output/img/smartit_login.png'")
    print("--- Fin de la prueba de RPA ---")

